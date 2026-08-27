from __future__ import annotations

import json
from pathlib import Path
import subprocess
import sys

import pandas as pd
from openpyxl import load_workbook

from ff_market_projections.runs import initialize_run
from ff_market_projections.workbook import SHEETS


ROOT = Path(__file__).parents[1]


def _write_inputs(run_dir: Path) -> None:
    artifacts = run_dir / "artifacts"
    pd.DataFrame([{
        "run_id": run_dir.name, "season": "2026-27", "canonical_player_id": "p1", "canonical_player_name": "An Extremely Long Player Name For Workbook Testing",
        "scoring_profile": "receiving", "components_missing": "", "projection_complete": True, "scoring_scope": "market_supported_stats_only",
        "receiving_yards": 1000.0, "receiving_touchdowns": 8.0, "receptions": 100.0,
        "partial_fpts_standard": 148.0, "partial_fpts_half_ppr": 198.0, "partial_fpts_three_quarter_ppr": 223.0, "partial_fpts_full_ppr": 248.0,
        "fpts_standard": 148.0, "fpts_half_ppr": 198.0, "fpts_three_quarter_ppr": 223.0, "fpts_full_ppr": 248.0,
    }]).to_csv(artifacts / "fantasy_projections.csv", index=False)
    pd.DataFrame([{"run_id": run_dir.name, "source": "draftkings", "canonical_player_id": "p1", "stat": "receiving_yards", "mean": 1000.0, "source_market_ids": "https://example.test/market"}]).to_csv(artifacts / "source_projections.csv", index=False)
    pd.DataFrame([{"source": "draftkings", "raw_player_name": "Long Name", "canonical_player_name": "An Extremely Long Player Name For Workbook Testing", "canonical_threshold": 999.5, "modeling_probability": 0.5, "inclusion_status": "included", "exclusion_reason": ""}]).to_csv(artifacts / "priced_markets.csv", index=False)
    pd.DataFrame([{"source": "draftkings", "raw_player_name": "Long Name", "canonical_player_id": "p1", "canonical_player_name": "An Extremely Long Player Name For Workbook Testing", "match_method": "exact", "match_score": 100.0, "review_status": "accepted"}]).to_csv(artifacts / "player_map.csv", index=False)
    pd.DataFrame([{"stat": "receiving_yards", "historical_dispersion": 5.0, "final_dispersion": 6.0, "method": "historical_plus_current_market_map", "status": "passed"}]).to_csv(artifacts / "dispersion_calibration.csv", index=False)
    (artifacts / "scoring_validation.json").write_text(json.dumps({"status": "passed", "checks": [{"name": "scoring.test", "passed": True, "severity": "error", "message": "fine", "details": {}}]}), encoding="utf-8")


def test_workbook_build_readback_and_static_values(tmp_path: Path) -> None:
    config = ROOT / "config/pipeline.toml"
    run_dir = initialize_run(__import__("ff_market_projections.config", fromlist=["load_config"]).load_config(config), ROOT / "config/player_aliases.csv", tmp_path / "runs")
    _write_inputs(run_dir)
    build = subprocess.run([sys.executable, str(ROOT / "scripts/build_workbook.py"), "--run-dir", str(run_dir)], cwd=ROOT, text=True, capture_output=True)
    assert build.returncode == 0, build.stderr
    validate = subprocess.run([sys.executable, str(ROOT / "scripts/validate_workbook.py"), "--run-dir", str(run_dir)], cwd=ROOT, text=True, capture_output=True)
    assert validate.returncode == 0, validate.stderr
    workbook = load_workbook(run_dir / "output/fantasy_football_projections.xlsx", data_only=False)
    assert workbook.sheetnames == list(SHEETS)
    assert all(workbook[name].freeze_panes == "A2" and workbook[name].tables for name in SHEETS)
    projections = workbook["Projections"]
    headers = [cell.value for cell in projections[1]]
    assert projections.cell(2, headers.index("fpts_full_ppr") + 1).value == 248.0
    assert all(not (isinstance(cell.value, str) and cell.value.startswith("=")) for row in projections.iter_rows() for cell in row)
    assert json.loads((run_dir / "artifacts/workbook_validation.json").read_text())["status"] == "passed"


def test_missing_artifact_fails_closed(tmp_path: Path) -> None:
    run_dir = tmp_path / "run"; (run_dir / "config").mkdir(parents=True); (run_dir / "output").mkdir(); (run_dir / "artifacts").mkdir()
    (run_dir / "config/effective.toml").write_bytes((ROOT / "config/pipeline.toml").read_bytes())
    failed = subprocess.run([sys.executable, str(ROOT / "scripts/build_workbook.py"), "--run-dir", str(run_dir)], cwd=ROOT, text=True, capture_output=True)
    assert failed.returncode != 0
    assert "Required workbook artifact is missing" in failed.stderr
