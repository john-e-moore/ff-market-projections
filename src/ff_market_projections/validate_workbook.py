"""Read-back checks for the static Excel deliverable."""

from __future__ import annotations

import math
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .workbook import SHEETS
from .contracts import sha256_file


def validate_workbook(run_dir: str | Path, path: str | Path) -> dict[str, Any]:
    run_dir, path = Path(run_dir), Path(path)
    workbook = load_workbook(path, data_only=False)
    checks: list[dict[str, Any]] = []
    def check(name: str, passed: bool, message: str, **details: Any) -> None:
        checks.append({"name": name, "passed": passed, "severity": "error", "message": message, "details": details})
    check("workbook.required_sheets", workbook.sheetnames == list(SHEETS), "All eight required sheets exist in contract order", actual=workbook.sheetnames)
    for name in SHEETS:
        sheet = workbook[name]
        check(f"workbook.{name}.table", bool(sheet.tables), "Worksheet has an Excel table")
        check(f"workbook.{name}.freeze_panes", sheet.freeze_panes == "A2", "Worksheet freezes its header row")
        headers = [cell.value for cell in sheet[1]]
        check(f"workbook.{name}.headers", len(headers) == len(set(headers)) and all(headers), "Worksheet headers are unique and populated")
    errors = {"#REF!", "#VALUE!", "#DIV/0!"}
    invalid = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in errors:
                    invalid.append(f"{sheet.title}!{cell.coordinate}")
                if isinstance(cell.value, float) and not math.isfinite(cell.value):
                    invalid.append(f"{sheet.title}!{cell.coordinate}")
    check("workbook.no_error_values", not invalid, "Workbook has no Excel errors, NaN, or infinity", cells=invalid)
    for sheet, filename in {"Projections": "fantasy_projections.csv", "Source Projections": "source_projections.csv", "Market Inputs": "priced_markets.csv", "Name Map": "player_map.csv", "Calibration": "dispersion_calibration.csv"}.items():
        expected = len(pd.read_csv(run_dir / "artifacts" / filename, low_memory=False))
        check(f"workbook.{sheet}.row_count", workbook[sheet].max_row - 1 == expected, "Workbook row count matches source artifact", expected=expected, observed=workbook[sheet].max_row - 1)
    projection = workbook["Projections"]
    headers = [cell.value for cell in projection[1]]
    if projection.max_row > 1 and "fpts_full_ppr" in headers:
        value = projection.cell(2, headers.index("fpts_full_ppr") + 1).value
        check("workbook.projection_numeric", value is None or isinstance(value, (int, float)), "Representative fantasy projection is numeric")
        source = pd.read_csv(run_dir / "artifacts" / "fantasy_projections.csv", low_memory=False).iloc[0]
        expected = source["fpts_full_ppr"]
        check("workbook.projection_reconciliation", (pd.isna(expected) and value is None) or abs(float(value) - float(expected)) <= 1e-9, "Sample fantasy value exactly matches CSV artifact", expected=None if pd.isna(expected) else float(expected), observed=value)
    source_projection = pd.read_csv(run_dir / "artifacts" / "source_projections.csv", low_memory=False)
    published = pd.read_csv(run_dir / "artifacts" / "fantasy_projections.csv", low_memory=False)
    source_players = set(source_projection.get("canonical_player_id", pd.Series(dtype=str)).dropna())
    published_players = set(published.get("canonical_player_id", pd.Series(dtype=str)).dropna())
    check("workbook.source_lineage", published_players <= source_players, "Every published player has source-projection lineage", missing=sorted(published_players - source_players))
    manifest_path = run_dir / "metadata" / "manifest.json"
    if manifest_path.is_file():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        config_hash = sha256_file(run_dir / "config" / "effective.toml")
        check("workbook.config_hash", manifest.get("effective_config_sha256") == config_hash, "Manifest configuration hash matches the run-scoped effective config", expected=config_hash, observed=manifest.get("effective_config_sha256"))
    workbook.save(path)
    return {"status": "passed" if all(check["passed"] for check in checks) else "failed", "checks": checks}
