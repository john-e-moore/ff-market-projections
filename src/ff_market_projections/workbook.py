"""Static, auditable Excel workbook construction."""

from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEETS = ("Projections", "Source Projections", "Market Inputs", "Name Map", "Calibration", "Scoring", "Validation", "Run Info")
CSV_SHEETS = {
    "Source Projections": "source_projections.csv",
    "Market Inputs": "priced_markets.csv",
    "Name Map": "player_map.csv",
    "Calibration": "dispersion_calibration.csv",
}
VALIDATION_FILES = ("collections_validation.json", "historical_validation.json", "normalized_validation.json", "identity_validation.json", "pricing_validation.json", "historical_calibration.json", "model_validation.json", "aggregation_validation.json", "scoring_validation.json")


class WorkbookError(ValueError):
    pass


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.is_file():
        raise WorkbookError(f"Required workbook artifact is missing: {path.name}")
    return pd.read_csv(path, low_memory=False)


def _json(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {"status": "not_available", "checks": []}
    return json.loads(path.read_text(encoding="utf-8"))


def _projections(fantasy: pd.DataFrame) -> pd.DataFrame:
    required = {"canonical_player_id", "canonical_player_name", "fpts_full_ppr", "projection_complete"}
    missing = sorted(required - set(fantasy.columns))
    if missing:
        raise WorkbookError(f"fantasy_projections.csv missing required columns: {', '.join(missing)}")
    preferred = ["canonical_player_id", "canonical_player_name", "fpts_full_ppr", "fpts_three_quarter_ppr", "fpts_half_ppr", "fpts_standard", "partial_fpts_full_ppr", "partial_fpts_three_quarter_ppr", "partial_fpts_half_ppr", "partial_fpts_standard", "projection_complete", "scoring_scope", "scoring_profile", "components_missing"]
    stats = [column for column in fantasy.columns if column in {"passing_yards", "passing_touchdowns", "passing_interceptions", "rushing_yards", "rushing_touchdowns", "receiving_yards", "receiving_touchdowns", "receptions", "fumbles_lost", "two_point_conversions"}]
    remaining = [column for column in fantasy.columns if column not in preferred + stats]
    return fantasy[[column for column in preferred + stats + remaining if column in fantasy.columns]].copy()


def _scoring(config: dict[str, Any]) -> pd.DataFrame:
    scoring = config["scoring"]
    rows = []
    for key, value in scoring.items():
        if key == "required_profiles":
            for profile, components in value.items():
                rows.append({"setting": f"required_profiles.{profile}", "value": "|".join(components), "supported_by_current_markets": all(component in {"passing_yards", "passing_touchdowns", "rushing_yards", "rushing_touchdowns", "receiving_yards", "receiving_touchdowns", "receptions"} for component in components)})
        elif key == "reception_bonus":
            for mode, bonus in value.items():
                rows.append({"setting": f"reception_bonus.{mode}", "value": bonus, "supported_by_current_markets": True})
        else:
            rows.append({"setting": key, "value": value, "supported_by_current_markets": key in {"missing_stat_policy", "passing_yards", "passing_touchdowns", "rushing_yards", "rushing_touchdowns", "receiving_yards", "receiving_touchdowns"}})
    return pd.DataFrame(rows)


def _validation(artifacts: Path) -> pd.DataFrame:
    rows = []
    for filename in VALIDATION_FILES:
        report = _json(artifacts / filename)
        for check in report.get("checks", []):
            rows.append({"task": filename.removesuffix("_validation.json").removesuffix(".json"), "check_name": check.get("name", ""), "severity": check.get("severity", "error"), "observed_value": json.dumps(check.get("details", {}), sort_keys=True), "expected_rule": check.get("message", ""), "passed": check.get("passed"), "message": check.get("message", "")})
    return pd.DataFrame(rows, columns=["task", "check_name", "severity", "observed_value", "expected_rule", "passed", "message"])


def _run_info(run_dir: Path, config: dict[str, Any]) -> pd.DataFrame:
    manifest = _json(run_dir / "metadata" / "manifest.json")
    environment = _json(run_dir / "metadata" / "environment.json")
    warnings = []
    for filename in VALIDATION_FILES:
        warnings.extend(check.get("message", "") for check in _json(run_dir / "artifacts" / filename).get("checks", []) if check.get("severity") == "warning" and not check.get("passed"))
    values = {
        "run_id": run_dir.name, "season": config["run"]["season"], "point_in_time_note": "Market-derived estimates reflect the source snapshot times recorded below.",
        "scoring_scope": "market_supported_stats_only", "dispersion_note": "A shared per-stat historical calibration is retained for audit; eligible multi-threshold current-market curves update or override misspecified historical shape as recorded in Calibration.",
        "source_snapshot_times_utc": "|".join(manifest.get("source_snapshot_times_utc", [])), "historical_provenance": json.dumps(manifest.get("historical", {}), sort_keys=True),
        "effective_config_sha256": environment.get("effective_config_sha256", manifest.get("effective_config_sha256", "")), "git_environment": json.dumps(environment, sort_keys=True),
        "warnings": " | ".join(warnings) if warnings else "None",
    }
    return pd.DataFrame([{"field": key, "value": value} for key, value in values.items()])


def _value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value.item() if hasattr(value, "item") else value


def _write_table(sheet, frame: pd.DataFrame, table_name: str) -> None:
    headers = list(frame.columns) or ["no_rows"]
    if frame.empty:
        frame = pd.DataFrame(columns=headers)
    sheet.append(headers)
    for row in frame.itertuples(index=False, name=None):
        sheet.append([_value(value) for value in row])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    end = max(2, sheet.max_row)
    sheet.add_table(Table(displayName=table_name, ref=f"A1:{get_column_letter(sheet.max_column)}{end}", tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)))
    sheet.freeze_panes = "A2"
    for index, header in enumerate(headers, 1):
        width = min(55, max(12, len(str(header)) + 2, max((len(str(sheet.cell(row, index).value or "")) for row in range(2, min(sheet.max_row, 25) + 1)), default=0) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook(run_dir: str | Path, config: dict[str, Any], output: str | Path) -> dict[str, int]:
    run_dir, output = Path(run_dir), Path(output)
    artifacts = run_dir / "artifacts"
    fantasy = _read_csv(artifacts / "fantasy_projections.csv")
    frames: dict[str, pd.DataFrame] = {"Projections": _projections(fantasy), "Scoring": _scoring(config), "Validation": _validation(artifacts), "Run Info": _run_info(run_dir, config)}
    frames.update({sheet: _read_csv(artifacts / filename) for sheet, filename in CSV_SHEETS.items()})
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, name in enumerate(SHEETS, 1):
        sheet = workbook.create_sheet(name)
        _write_table(sheet, frames[name], f"Table{index}")
        if name == "Projections":
            if "projection_complete" in frames[name]:
                column = get_column_letter(list(frames[name].columns).index("projection_complete") + 1)
                sheet.conditional_formatting.add(f"{column}2:{column}{sheet.max_row}", FormulaRule(formula=[f"{column}2=FALSE"], fill=PatternFill("solid", fgColor="FFF2CC")))
        if name == "Validation" and "passed" in frames[name]:
            column = get_column_letter(list(frames[name].columns).index("passed") + 1)
            sheet.conditional_formatting.add(f"{column}2:{column}{sheet.max_row}", FormulaRule(formula=[f"{column}2=FALSE"], fill=PatternFill("solid", fgColor="F4CCCC")))
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {name: len(frame) for name, frame in frames.items()}
